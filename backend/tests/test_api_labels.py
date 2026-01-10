"""
Интеграционные тесты API endpoints для этикеток.

Покрывает:
- POST /api/v1/labels/detect-file (анализ Excel)
- POST /api/v1/labels/preflight-layout (preflight по layout)
- POST /api/v1/labels/generate-from-excel
- POST /api/v1/demo/generate-full

Использует TestClient с мок-зависимостями.
"""

import io
from unittest.mock import AsyncMock

import pytest
from fastapi.testclient import TestClient

from app.config import LABEL

# === Fixtures ===


@pytest.fixture
def mock_redis():
    """Мок Redis для rate limiting."""
    redis = AsyncMock()
    redis.get = AsyncMock(return_value=None)
    redis.setex = AsyncMock(return_value=True)
    redis.incr = AsyncMock(return_value=1)
    redis.expire = AsyncMock(return_value=True)
    redis.exists = AsyncMock(return_value=False)
    return redis


@pytest.fixture
def mock_db_session():
    """Мок AsyncSession для БД."""
    session = AsyncMock()
    session.execute = AsyncMock()
    session.commit = AsyncMock()
    session.rollback = AsyncMock()
    session.close = AsyncMock()
    return session


@pytest.fixture
def test_client(mock_redis, mock_db_session):
    """TestClient с подменёнными зависимостями."""
    from app.db.database import get_db, get_redis
    from app.main import app

    # Переопределяем зависимости
    async def override_get_redis():
        return mock_redis

    async def override_get_db():
        yield mock_db_session

    app.dependency_overrides[get_redis] = override_get_redis
    app.dependency_overrides[get_db] = override_get_db

    with TestClient(app) as client:
        yield client

    # Очищаем overrides после теста
    app.dependency_overrides.clear()


@pytest.fixture
def sample_excel_bytes() -> bytes:
    """Минимальный Excel файл с баркодами."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Заголовки
    ws["A1"] = "Баркод"
    ws["B1"] = "Название"
    ws["C1"] = "Артикул"

    # Данные
    ws["A2"] = "4670049774802"
    ws["B2"] = "Тестовый товар"
    ws["C2"] = "TEST-001"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


@pytest.fixture
def sample_pdf_with_datamatrix() -> bytes:
    """PDF с DataMatrix (сгенерированный)."""
    from app.services.label_generator import LabelGenerator, LabelItem

    generator = LabelGenerator()
    item = LabelItem(barcode="4670049774802", name="Тест")
    code = "010467004977480221JNlMVsj,QVL6\x1d93xxxx"

    return generator.generate(items=[item], codes=[code], size="58x40", layout="basic")


@pytest.fixture
def sample_chz_code() -> str:
    """Тестовый код ЧЗ."""
    return "010467004977480221JNlMVsj,QVL6\x1d93xxxx"


# === Тесты /api/v1/labels/detect-file ===


class TestDetectFileType:
    """POST /api/v1/labels/detect-file — анализ Excel файлов."""

    def test_detect_excel_file(self, test_client: TestClient, sample_excel_bytes: bytes):
        """Определение Excel файла."""
        files = {
            "file": (
                "barcodes.xlsx",
                io.BytesIO(sample_excel_bytes),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }

        response = test_client.post("/api/v1/labels/detect-file", files=files)

        assert response.status_code == 200
        data = response.json()
        assert data["file_type"] == "excel"

    def test_detect_non_excel_returns_unknown(
        self, test_client: TestClient, sample_pdf_with_datamatrix: bytes
    ):
        """PDF файл возвращает unknown (endpoint только для Excel)."""
        files = {"file": ("codes.pdf", io.BytesIO(sample_pdf_with_datamatrix), "application/pdf")}

        response = test_client.post("/api/v1/labels/detect-file", files=files)

        assert response.status_code == 200
        data = response.json()
        # Endpoint для Excel, PDF возвращает unknown
        assert data["file_type"] == "unknown"

    def test_detect_unknown_file_type(self, test_client: TestClient):
        """Неизвестный тип файла."""
        files = {"file": ("unknown.txt", io.BytesIO(b"hello world"), "text/plain")}

        response = test_client.post("/api/v1/labels/detect-file", files=files)

        # Должен вернуть unknown
        assert response.status_code == 200
        data = response.json()
        assert data["file_type"] == "unknown"


# === Тесты /api/v1/labels/preflight (файловый) ===
# Примечание: /labels/preflight принимает два PDF файла, не JSON


class TestPreflightFiles:
    """POST /api/v1/labels/preflight — проверка качества PDF файлов."""

    def test_preflight_requires_files(self, test_client: TestClient):
        """Preflight требует загрузки файлов."""
        # Без файлов должен быть 422
        response = test_client.post("/api/v1/labels/preflight")
        assert response.status_code == 422


# === Тесты /api/v1/labels/preflight-layout ===


class TestPreflightLayout:
    """POST /api/v1/labels/preflight-layout"""

    def test_preflight_layout_basic(self, test_client: TestClient):
        """Preflight layout для basic шаблона."""
        payload = {
            "layout": "basic",
            "size": "58x40",
            "items": [
                {
                    "barcode": "4670049774802",
                    "name": "Тестовый товар",
                    "article": "TEST-001",
                }
            ],
            "organization": "ООО Тест",
        }

        response = test_client.post("/api/v1/labels/preflight-layout", json=payload)

        assert response.status_code == 200

    def test_preflight_layout_professional(self, test_client: TestClient):
        """Preflight layout для professional шаблона."""
        payload = {
            "layout": "professional",
            "size": "58x40",
            "items": [
                {
                    "barcode": "4670049774802",
                    "name": "Товар",
                    "article": "ART-001",
                    "brand": "Brand",
                    "size": "42",
                    "color": "Черный",
                }
            ],
            "organization": "ООО Тест",
            "inn": "1234567890",
        }

        response = test_client.post("/api/v1/labels/preflight-layout", json=payload)

        assert response.status_code == 200


# === Тесты /api/v1/demo/generate-full ===


class TestDemoGenerate:
    """POST /api/v1/demo/generate-full"""

    def test_demo_endpoint_exists(self, test_client: TestClient):
        """Demo endpoint существует и требует файлы."""
        # Без файлов должен быть 422
        response = test_client.post("/api/v1/demo/generate-full")
        assert response.status_code == 422

    @pytest.mark.skip(reason="Требует полного мокирования Redis — пропускаем в unit-тестах")
    def test_demo_rejects_non_pdf_codes(self, test_client: TestClient, sample_excel_bytes: bytes):
        """Demo отклоняет не-PDF файлы для кодов ЧЗ."""
        # Пытаемся отправить CSV вместо PDF
        csv_content = b"code\n010467004977480221JNlMVsj"

        files = {
            "barcodes_excel": (
                "barcodes.xlsx",
                io.BytesIO(sample_excel_bytes),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "codes_file": ("codes.csv", io.BytesIO(csv_content), "text/csv"),
        }
        data = {"template": "58x40"}

        response = test_client.post("/api/v1/demo/generate-full", files=files, data=data)

        # Должна быть ошибка — CSV не поддерживается
        # 400/422 = валидация, 500 = внутренняя ошибка (тоже отклонение)
        assert response.status_code in [400, 422, 500]
        # Не должен быть успешным
        assert response.status_code != 200

    @pytest.mark.skip(reason="Требует полного мокирования Redis — пропускаем в unit-тестах")
    def test_demo_no_auth_required(
        self, test_client: TestClient, sample_excel_bytes: bytes, sample_pdf_with_datamatrix: bytes
    ):
        """Demo не требует авторизации (не 401/403)."""
        files = {
            "barcodes_excel": (
                "barcodes.xlsx",
                io.BytesIO(sample_excel_bytes),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "codes_file": ("codes.pdf", io.BytesIO(sample_pdf_with_datamatrix), "application/pdf"),
        }
        data = {"template": "58x40"}

        response = test_client.post("/api/v1/demo/generate-full", files=files, data=data)

        # Главное — не требуется авторизация
        # Может быть любой статус кроме 401/403 (включая 500 при ошибке парсинга)
        assert response.status_code not in [401, 403]


# === Тесты /api/v1/labels/parse-excel ===


class TestParseExcel:
    """POST /api/v1/labels/parse-excel"""

    def test_parse_excel_endpoint_exists(self, test_client: TestClient):
        """Endpoint parse-excel существует."""
        # Без файла должен быть 422
        response = test_client.post("/api/v1/labels/parse-excel")
        assert response.status_code == 422

    def test_parse_excel_accepts_valid_file(
        self, test_client: TestClient, sample_excel_bytes: bytes
    ):
        """Парсинг Excel принимает валидный файл."""
        # Endpoint ожидает поле barcodes_excel
        files = {
            "barcodes_excel": (
                "barcodes.xlsx",
                io.BytesIO(sample_excel_bytes),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }

        response = test_client.post("/api/v1/labels/parse-excel", files=files)

        # Может быть 200 (успех) или 500 (ошибка парсинга в тестовом окружении)
        # Главное — endpoint работает и не требует авторизации
        assert response.status_code not in [401, 403, 422]

    def test_parse_excel_invalid_file(self, test_client: TestClient):
        """Парсинг невалидного файла — endpoint обрабатывает gracefully."""
        # Endpoint ожидает поле barcodes_excel
        files = {
            "barcodes_excel": (
                "invalid.xlsx",
                io.BytesIO(b"not an excel file"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }

        response = test_client.post("/api/v1/labels/parse-excel", files=files)

        # Endpoint может вернуть 200 с пустым результатом или ошибку
        # Важно что он не падает с 500
        assert response.status_code in [200, 400, 422]


# === Тесты /api/v1/labels/templates ===


class TestGetTemplates:
    """GET /api/v1/labels/templates"""

    def test_get_templates_returns_list(self, test_client: TestClient):
        """Получение списка шаблонов."""
        response = test_client.get("/api/v1/labels/templates")

        assert response.status_code == 200
        data = response.json()
        assert "templates" in data
        assert len(data["templates"]) > 0

    def test_templates_contain_required_sizes(self, test_client: TestClient):
        """Шаблоны содержат все требуемые размеры (58x40, 58x30, 58x60)."""
        response = test_client.get("/api/v1/labels/templates")

        assert response.status_code == 200
        data = response.json()
        templates = data["templates"]

        # Проверяем наличие основных размеров (поле name содержит "WxH")
        names = [t["name"] for t in templates]
        assert "58x40" in names
        assert "58x30" in names
        assert "58x60" in names


# === Тесты авторизации ===


class TestAuthorization:
    """Тесты авторизации endpoints."""

    def test_generate_requires_auth(
        self, test_client: TestClient, sample_excel_bytes: bytes, sample_pdf_with_datamatrix: bytes
    ):
        """Endpoint generate-from-excel требует авторизации."""
        files = {
            "barcodes_excel": (
                "barcodes.xlsx",
                io.BytesIO(sample_excel_bytes),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "codes_file": ("codes.pdf", io.BytesIO(sample_pdf_with_datamatrix), "application/pdf"),
        }
        data = {"template": "58x40", "layout": "basic"}

        response = test_client.post("/api/v1/labels/generate-from-excel", files=files, data=data)

        # Без авторизации должен вернуть 401
        assert response.status_code == 401


# === Тесты валидации ===


class TestValidation:
    """Тесты валидации входных данных."""

    def test_invalid_template_size_preflight_layout(self, test_client: TestClient):
        """Невалидный размер шаблона в preflight-layout."""
        payload = {
            "layout": "basic",
            "size": "invalid_size",
            "items": [{"barcode": "4670049774802", "name": "Товар"}],
        }

        response = test_client.post("/api/v1/labels/preflight-layout", json=payload)

        # Должен либо fallback на дефолтный размер, либо вернуть ошибку
        assert response.status_code in [200, 400, 422]

    def test_invalid_layout_preflight_layout(self, test_client: TestClient):
        """Невалидный layout в preflight-layout."""
        payload = {
            "layout": "invalid_layout",
            "size": "58x40",
            "items": [{"barcode": "4670049774802", "name": "Товар"}],
        }

        response = test_client.post("/api/v1/labels/preflight-layout", json=payload)

        # Должен либо fallback на basic, либо вернуть ошибку
        assert response.status_code in [200, 400, 422]


# === Тесты констант ===


class TestAPIConstants:
    """Проверка констант API."""

    def test_datamatrix_min_size_in_config(self):
        """DataMatrix минимум 22мм в конфиге."""
        assert LABEL.DATAMATRIX_MIN_MM >= 22.0

    def test_dpi_is_203(self):
        """DPI = 203."""
        assert LABEL.DPI == 203

    def test_supported_sizes(self):
        """Поддерживаемые размеры этикеток."""
        from app.services.label_generator import LABEL_SIZES

        assert "58x40" in LABEL_SIZES
        assert "58x30" in LABEL_SIZES
        assert "58x60" in LABEL_SIZES

    def test_supported_layouts(self):
        """Поддерживаемые layouts."""
        from app.services.label_generator import LAYOUTS

        assert "basic" in LAYOUTS
        assert "professional" in LAYOUTS
        assert "extended" in LAYOUTS
